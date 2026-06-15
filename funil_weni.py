#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Funil Comercial Weni x Contatos x Pedidos -- Amara NZero
========================================================
Le tres fontes e gera o painel HTML + a planilha XLSX:
  - conversas  (message_export*.xlsx)  -- exportacao da Weni / API
  - contatos   (contact_export*.xlsx)  -- base com vendedor, CNPJ, UF
  - pedidos    (Relatorio*.xlsx / pedidos*.xlsx) -- relatorio de pedidos (cruzado por CNPJ)

USO:
    python funil_weni.py --input entradas --output saidas

Procura na pasta --input:
  message_export*.xlsx, contact_export*.xlsx, e um arquivo de pedidos
  (nome contendo "relatorio", "pedido" ou "order"). O de pedidos e opcional:
  sem ele, "Fecharam pedido" fica vazio (nao da pra confirmar por CNPJ).
"""
import argparse, collections, glob, json, os, re, sys, unicodedata
from datetime import datetime, timedelta

try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError as e:
    sys.exit("Dependencia faltando: %s\nRode:  pip install -r requirements.txt" % e)

PRODUCTION_CHANNEL = "Amara NZero"

# ---------------------------------------------------------------------------
# Contatos a EXCLUIR da analise (equipe interna, vendedores como atendentes,
# contas de teste). Edite esta lista conforme necessario.
# ---------------------------------------------------------------------------
EXCLUDE_NAMES = [
    "Maria Dantas V. Ferreira", "Maria Dantas", "Camila Dias", "M\u00f4nica Silva",
    "Amanda Barbosa", "Bruno barreto", "Kalila Caetano", "Andressa Silva",
    "Marcelo Souza", "Rafaela Menezes", "Rainei Trindade", "Vanessa Vieira",
    "Virginia Vieira", "Virg\u00ednia Vieira", "Luana Castilho",
    "Luana Castilho Comercial Amara", "Gabriel Borges", "Elen Cruz",
    "Celen\u00ea Carmo", "Artur - Suporte T\u00e9cnico", "Augusto Batista",
    "Produtos Financeiros", "Marketing", "SAC - Amara", "- Amara NZero",
]


def _norm_name(s):
    s = str(s).lower().strip()
    s = "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")
    return re.sub(r"\s+", " ", s)


_EXCLUDE_SET = {_norm_name(x) for x in EXCLUDE_NAMES}


def is_excluded(name):
    if not isinstance(name, str):
        return False
    n = _norm_name(name)
    if n in _EXCLUDE_SET:
        return True
    if "amara nzero |" in n or "| amara nzero" in n:   # contas de atendente
        return True
    if n.endswith("- amara") or n.endswith("-amara") or n.endswith("| amara"):
        return True
    if "@amaranzero.com" in n or "@vtex.com" in n or "suporte weni" in n:
        return True
    return False

AUTO_PAT = re.compile(r"(?:agradecemos (?:o |seu )?contato|fora do (?:nosso )?hor\u00e1rio|hor\u00e1rio de atendimento|no momento[, ]+n\u00e3o estamos|resposta autom\u00e1tica|atendimento autom\u00e1tico|chave pix|deixe sua mensagem|retornaremos (?:o |seu )?contato|assim que poss\u00edvel retorn|seja bem[- ]vindo ao|voc\u00ea est\u00e1 falando com|menu principal|digite o n\u00famero)", re.I)
AGENT_PREFIX = re.compile(r"^([A-Z\u00c0-\u00dd][a-z\u00e0-\u00fd\u00e7]+):\s*\n")
BOT_PHRASE = re.compile(r"como voc\u00ea avalia|deixe um coment\u00e1rio|atendimento foi finalizado|assistente virtual|enviei um c\u00f3digo|percebi que talvez|agradecemos a sua colabora|sou a assistente", re.I)
FLUIG_CLIENT = re.compile(r"\b(fluig)\b|repasse (?:da nota|do pedido|de nota|fiscal)|fazer (?:o )?repasse|preciso (?:do |de )?repasse|sobre (?:o )?repasse", re.I)
FLOW_ANSWERS = re.compile(r"^(sim|n\u00e3o|ok|1|2|3|4|5|topo|topa|claro|pode ser|oi|ol\u00e1|bom dia|boa tarde|boa noite|obrigad[oa]|valeu|blz|beleza|ola)[\s!.,]*$", re.I)
ELSEWHERE_NEG = re.compile(r"n\u00e3o comprei|nunca comprei", re.I)
AGENT_BAD = re.compile(r"programa de pontos|repasse na amara|demonstra\u00e7\u00e3o conta e ordem|voc\u00ea (?:quer|j\u00e1 tem)|vai emitir depois|podemos tentar", re.I)

CLIENT_CLOSE = [
 (r"pedido\s*(?:n[\u00ba\u00b0o\.]?|numero|n\u00famero)?\s*#?\d{4,}", "Citou n\u00famero de pedido"),
 (r"\b(?:fechei|finalizei|conclu\u00ed|fechamos|finalizamos)\b[^.!?\n]{0,40}\b(?:o |meu |a |minha )?(?:pedido|compra)\b", "Confirmou ter fechado o pedido"),
 (r"\b(?:fiz|realizei|efetuei|acabei de fazer)\b[^.!?\n]{0,30}\b(?:o |um |uma |minha |meu )?(?:pedido|compra)\b", "Confirmou ter feito o pedido"),
 (r"\bcomprovante\b[^.!?\n]{0,40}(?:pagamento|pgto|compra|pedido|boleto|pix)|segue (?:o )?comprovante|comprovante (?:de|do) (?:pgto|pagamento)", "Enviou comprovante de pagamento"),
 (r"\b(?:j\u00e1 )?paguei\b|pagamento (?:foi )?(?:feito|realizado|efetuado|aprovado)|fiz o pagamento|fiz o pix", "Confirmou pagamento"),
 (r"\b(?:meu |o |do )pedido\b[^.!?\n]{0,40}\b(?:chegou|foi faturado|foi entregue|saiu para entrega)\b|\brastreio (?:do|de) (?:meu )?pedido\b", "Pedido j\u00e1 faturado/entregue"),
 (r"\bvou subir o pedido\b|\bsubi o pedido\b|\bpode faturar\b|pedido que acabamos de fechar|pedido (?:j\u00e1 )?(?:foi )?confirmado pela amara", "Autorizou faturar / pedido confirmado"),
]
AGENT_CLOSE = [
 (r"pagamento confirmado com sucesso", "Atendente: pagamento confirmado"),
 (r"seu pedido (?:est\u00e1|foi) (?:confirmado|faturado|em separa\u00e7\u00e3o|em processamento)", "Atendente: pedido confirmado/em separa\u00e7\u00e3o"),
 (r"pedido \*?confirmado\*?", "Atendente: pedido confirmado"),
 (r"retirada est\u00e1 prevista", "Atendente: retirada agendada"),
 (r"nota (?:fiscal )?(?:foi )?emitida com sucesso", "Atendente: nota emitida"),
]
AGENT_PROGRESS = [
 (r"envie?\s*(?:o\s*)?(?:n\u00famero\s*d[eo]\s*)?pedido|n\u00famero do (?:seu )?pedido", "Atendente pediu n\u00ba do pedido"),
 (r"(?:envie?|anexe?|manda?r?)\s*(?:o\s*)?comprovante|ao finalizar (?:pode )?envi", "Atendente pediu comprovante"),
 (r"(?:pode|consegue) finalizar (?:seu |o )?pedido|finalizar (?:a )?compra (?:na plataforma|no site)", "Atendente orientou finalizar"),
]
CONTATAR = [
 (r"\b(?:fechar|comprar|finalizar)\b[^.!?\n]{0,15}\bagora\b", "Quer fechar/comprar agora", 5),
 (r"\b(?:quero|vou|pretendo|gostaria de|preciso|posso)\b[^.!?\n]{0,30}\b(?:fechar|comprar|finalizar|adquirir|levar|pegar)\b", "Manifestou inten\u00e7\u00e3o de fechar", 4),
 (r"\bdesconto\b|\bnegoci\w+|\bmelhor pre\u00e7o\b|\bcondi\u00e7\u00e3o (?:de pagamento|especial|melhor)\b|\babatimento\b", "Pediu desconto / negociou", 4),
 (r"\bcupom\b", "Falou de cupom", 3),
 (r"\b(?:como (?:fa\u00e7o|posso) (?:para )?comprar|forma de pagamento|como pago|fechar o or\u00e7amento)\b", "Perguntou como comprar/pagar", 3),
 (r"\b(?:fechar|comprar)\b[^.!?\n]{0,20}\b(?:hoje|essa semana|ainda hoje|amanh\u00e3)\b", "Quer fechar em breve", 5),
]
ORCAMENTO = [(r"\b(?:or\u00e7amento|or\u00e7ar|or\u00e7a)\b", "Mencionou or\u00e7amento"), (r"\b(?:cota\u00e7\u00e3o|cotar|me cota|cotando)\b", "Pediu cota\u00e7\u00e3o")]
PERDIDO = [
 (r"comprei\s[^.!?\n]{0,30}(?:na weg|outra empresa|concorrente|outro fornecedor|outro lugar|com a aldo|na aldo|na edeltec)|fechei com (?:outra|outro)", "Comprou no concorrente"),
 (r"\b(?:muito caro|t\u00e1 caro|est\u00e1 caro|ficou caro|pre\u00e7o (?:muito )?(?:alto|elevado)|acima do mercado|mais caro que|valor (?:muito )?alto)\b", "Achou caro / condi\u00e7\u00e3o comercial"),
 (r"\b(?:sem estoque|n\u00e3o tem (?:em )?estoque|fora de estoque|produto indispon\u00edvel)\b", "Sem estoque"),
 (r"\b(?:n\u00e3o tenho (?:mais )?interesse|sem interesse|n\u00e3o quero mais|desisti|deixa pra l\u00e1|j\u00e1 resolvi|n\u00e3o vou (?:comprar|fechar)|cancelar? (?:o )?pedido)\b", "Recusou / desistiu"),
 (r"\b(?:prazo (?:muito )?(?:longo|grande|alto)|demora (?:muito|demais)|entrega (?:muito )?demorada)\b", "Prazo de entrega"),
 (r"\b(?:burocr\u00e1tic|muito complicado|dif\u00edcil (?:de )?comprar)\b", "Burocracia / fric\u00e7\u00e3o"),
]
INTENT = [
 (r"\bquero (?:fechar|seguir com|finalizar)\b|\bpodemos fechar\b|fechar (?:o )?(?:pedido|neg\u00f3cio)|seguir com (?:a )?compra|pr\u00f3ximo passo (?:para|pra) fechar", 5, "Quer fechar neg\u00f3cio"),
 (r"\bquero comprar agora\b|comprar agora|\bj\u00e1 quero (?:fazer|fechar|comprar)\b|fazer o pagamento|pode (?:gerar|emitir) (?:a )?(?:nota|nf|nota fiscal)|emiss\u00e3o (?:de|da) nota", 5, "Compra imediata / pagamento / NF"),
 (r"\b(?:tem|teria|tem algum)\b[^.!?\n]{0,15}desconto|consegue (?:melhorar|fazer melhor)|melhor(?:ar)? (?:o )?pre\u00e7o|\bdesconto\b|condi\u00e7\u00e3o (?:especial|melhor|de pagamento)", 4, "Desconto / negocia\u00e7\u00e3o"),
 (r"or\u00e7amento (?:para|pra) (?:fechar|comprar|efetivar|pedido)|quero (?:um )?or\u00e7amento", 4, "Or\u00e7amento para efetivar"),
 (r"pode (?:gerar|fazer|enviar) (?:a )?proposta|proposta comercial", 4, "Solicitou proposta"),
 (r"como (?:fa\u00e7o|faz|posso) (?:para |pra )?(?:realizar|fazer|fechar) (?:o )?(?:pedido|compra)|qual (?:o )?pr\u00f3ximo passo", 3, "Como proceder"),
]
TIME_PHRASES = [
 (r"\b(?:daqui|em)\s*(\d{1,2})\s*dias?\b", lambda m: ("days", int(m.group(1)))),
 (r"\bdepois de amanh\u00e3\b", lambda m: ("days", 2)),
 (r"\bamanh\u00e3\b", lambda m: ("days", 1)),
 (r"\b(?:semana que vem|pr\u00f3xima semana)\b", lambda m: ("days", 7)),
 (r"\b(?:segunda|segunda-feira)\b", lambda m: ("weekday", 0)),
 (r"\b(?:ter\u00e7a|ter\u00e7a-feira)\b", lambda m: ("weekday", 1)),
 (r"\b(?:quarta|quarta-feira)\b", lambda m: ("weekday", 2)),
 (r"\b(?:quinta|quinta-feira)\b", lambda m: ("weekday", 3)),
 (r"\b(?:sexta|sexta-feira)\b", lambda m: ("weekday", 4)),
]
CC=[(re.compile(p,re.I),d) for p,d in CLIENT_CLOSE]
ACR=[(re.compile(p,re.I),d) for p,d in AGENT_CLOSE]
AP=[(re.compile(p,re.I),d) for p,d in AGENT_PROGRESS]
CT=[(re.compile(p,re.I),d,w) for p,d,w in CONTATAR]
ORG=[(re.compile(p,re.I),d) for p,d in ORCAMENTO]
PE=[(re.compile(p,re.I),d) for p,d in PERDIDO]
INT=[(re.compile(p,re.I),s,d) for p,s,d in INTENT]
TP=[(re.compile(p,re.I),f) for p,f in TIME_PHRASES]

def norm_cnpj(v):
    if v is None or (isinstance(v,float)): 
        try:
            import math
            if isinstance(v,float) and math.isnan(v): return ""
        except: pass
    s=re.sub(r"\D","",str(v))
    return s.zfill(14) if s else ""

def clean(v):
    if v is None: return ""
    s=str(v).strip()
    return "" if s in ("","nan","none","NaN",".","S/ VENDEDOR","SEM EXECUTIVO") else s

def add_bd(s,d):
    if d<=0: return s
    x=s; a=0
    while a<d:
        x+=timedelta(days=1)
        if x.weekday()<5: a+=1
    return x

def next_wd(s,wd):
    x=s+timedelta(days=1)
    while x.weekday()!=wd: x+=timedelta(days=1)
    return x

def find_file(d,*prefixes):
    for pre in prefixes:
        hits=sorted(glob.glob(os.path.join(d,pre)),key=os.path.getmtime,reverse=True)
        if hits: return hits[0]
    return None

def load_orders(path):
    if not path: return {}
    ped=pd.read_excel(path)
    if "Status" in ped.columns:
        ped=ped[~ped["Status"].isin(["Cancelada"])]
    if "Dt. Cria\u00e7\u00e3o" in ped.columns:
        ped=ped.copy(); ped["_dt"]=pd.to_datetime(ped["Dt. Cria\u00e7\u00e3o"],errors="coerce")
        ped=ped.sort_values("_dt")
    out={}
    for _,r in ped.iterrows():
        cn=norm_cnpj(r.get("CNPJ"))
        if not cn: continue
        dt=r.get("_dt")
        out[cn]=dict(pedido=str(r.get("Pedido","")),status=str(r.get("Status","")),
                     data=dt.strftime("%d/%m/%Y") if pd.notna(dt) else "",
                     data_iso=dt.strftime("%Y-%m-%d") if pd.notna(dt) else "")
    return out

def load_contacts(path):
    ct=pd.read_excel(path,dtype=str)
    m={}
    for _,row in ct.iterrows():
        razao=clean(row.get("Field:Razao Social")) or clean(row.get("Field:Nome Fantasia"))
        cn=norm_cnpj(row.get("Field:CNPJ"))
        cidade=clean(row.get("Field:Cidade"))
        if cidade and cidade.isupper(): cidade=cidade.title()
        m[row["Contact UUID"]]=dict(
            vendedor=clean(row.get("Field:Analista")) or "Sem vendedor",
            cnpj=("%s.%s.%s/%s-%s"%(cn[:2],cn[2:5],cn[5:8],cn[8:12],cn[12:14])) if len(cn)==14 else "",
            cnpj_raw=cn, empresa=razao, regional=clean(row.get("Field:Regional")),
            uf=clean(row.get("Field:UF")) or clean(row.get("Field:Estado")), cidade=cidade)
    return m

def classify(conv_path, ct_map, orders):
    df=pd.read_excel(conv_path); df["Date"]=pd.to_datetime(df["Date"])
    df=df[df["Channel"].str.contains(PRODUCTION_CHANNEL,na=False)]
    period_start=df["Date"].min().strftime("%Y-%m-%d") if len(df) else "0000-00-00"
    first_dir=df.sort_values("Date").groupby("Contact UUID")["Direction"].first()
    last_info={}
    for uuid,g in df.groupby("Contact UUID"):
        g=g.sort_values("Date"); txt=g["Text"].iloc[-1]
        last_info[uuid]=dict(last_dir=g["Direction"].iloc[-1],last_date=g["Date"].iloc[-1],
                             last_text=str(txt)[:160] if pd.notna(txt) else "")
    real_in=df[(df["Direction"]=="IN") & df["Text"].notna()].copy()
    real_in=real_in[~real_in["Text"].str.contains(AUTO_PAT)]
    ins_uuids=set(real_in["Contact UUID"].unique())
    contacts=[]
    for uuid in ins_uuids:
        g=df[df["Contact UUID"]==uuid].sort_values("Date")
        _nm=g["Name"].iloc[0]
        if is_excluded(_nm):
            continue
        in_msgs=[(r["Date"],str(r["Text"])) for _,r in g.iterrows() if r["Direction"]=="IN" and pd.notna(r["Text"]) and not AUTO_PAT.search(str(r["Text"]))]
        out_msgs=[(r["Date"],str(r["Text"])) for _,r in g.iterrows() if r["Direction"]=="OUT" and pd.notna(r["Text"])]
        hit={"CLOSE":[],"ACLOSE":[],"PROG":[],"CONTATAR":[],"ORCAMENTO":[],"PERDIDO":[]}
        ctw=0; deadline=None; iscore=0; idesc=""
        for dt,t in in_msgs:
            for pat,desc in CC:
                if pat.search(t): hit["CLOSE"].append((desc,dt,t[:220].strip())); break
            for pat,desc in PE:
                if pat.search(t):
                    if desc=="Comprou no concorrente" and ELSEWHERE_NEG.search(t): continue
                    hit["PERDIDO"].append((desc,dt,t[:220].strip())); break
            for pat,desc,w in CT:
                if pat.search(t): hit["CONTATAR"].append((desc,dt,t[:220].strip())); ctw=max(ctw,w); break
            for pat,desc in ORG:
                if pat.search(t): hit["ORCAMENTO"].append((desc,dt,t[:220].strip())); break
            for pat,fn in TP:
                mm=pat.search(t)
                if mm: deadline=(fn(mm),dt,t[:120].strip()); break
            for pat,s,desc in INT:
                if pat.search(t) and s>iscore: iscore=s; idesc=desc
        for dt,t in out_msgs:
            if "?" not in t.strip()[-60:] and not AGENT_BAD.search(t):
                for pat,desc in ACR:
                    if pat.search(t): hit["ACLOSE"].append((desc,dt,t[:220].strip())); break
            for pat,desc in AP:
                if pat.search(t): hit["PROG"].append((desc,dt,t[:220].strip())); break
        # real human agent contact?
        human=[t for _,t in out_msgs if AGENT_PREFIX.match(t) and not BOT_PHRASE.search(t)]
        seller_replied=len(human)>0
        name=g["Name"].iloc[0]
        name=name if isinstance(name,str) and name.strip() not in ("",".","$","&") else "(sem nome)"
        meta=ct_map.get(uuid,{}); li=last_info[uuid]
        client_waiting=li["last_dir"]=="IN" and not FLOW_ANSWERS.match(li["last_text"].strip())
        convo_is_fluig=any(FLUIG_CLIENT.search(t) for _,t in in_msgs)
        cn=meta.get("cnpj_raw","")
        oi=orders.get(cn,{}); has_order=bool(oi)
        # FECHADO = a conversa indica fechamento (independente de CNPJ).
        # fluig/repasse sozinho nao fecha, mas um fechamento explicito do cliente vale.
        convo_closed=bool(hit["CLOSE"]) or bool(hit["ACLOSE"])
        if convo_is_fluig and not hit["CLOSE"]: convo_closed=False
        if convo_closed:
            stage="FECHADO"; ev=hit["CLOSE"][0] if hit["CLOSE"] else hit["ACLOSE"][0]
            ck="Cliente confirmou" if hit["CLOSE"] else "Atendente confirmou"
        elif hit["PERDIDO"]:
            stage,ev,ck="PERDIDO",hit["PERDIDO"][0],""
        elif (hit["CONTATAR"] or hit["PROG"]) and seller_replied:
            stage,ck="ENTROU",""; ev=max(hit["CONTATAR"],key=lambda e:e[1]) if hit["CONTATAR"] else hit["PROG"][-1]
        elif hit["CONTATAR"] or hit["PROG"]:
            stage,ck="CONTATAR",""; ev=max(hit["CONTATAR"],key=lambda e:e[1]) if hit["CONTATAR"] else hit["PROG"][-1]
            if not hit["CONTATAR"]: ctw=max(ctw,4)
        elif hit["ORCAMENTO"]:
            stage,ev,ck="ORCAMENTO",hit["ORCAMENTO"][-1],""
        else:
            stage,ev,ck="SEM_SINAL",None,""
        contact_date=None; basis=""
        if stage=="CONTATAR":
            last_d=li["last_date"].normalize()
            if deadline:
                (kind,val),bdt,btxt=deadline; base=bdt.normalize()
                contact_date=add_bd(base,val) if kind=="days" else next_wd(base,val); basis='Cliente disse: "%s"'%btxt
            else:
                if ctw>=5 or client_waiting: contact_date=add_bd(last_d,0 if client_waiting else 1); tier="quente"
                elif ctw>=4: contact_date=add_bd(last_d,2); tier="morno"
                else: contact_date=add_bd(last_d,5); tier="frio"
                basis="Cad\u00eancia %s (sem prazo dito)"%tier
        contacts.append(dict(uuid=uuid,name=name,urn=str(g["URN"].iloc[0]),
            vendedor=meta.get("vendedor","Sem vendedor"),cnpj=meta.get("cnpj",""),empresa=meta.get("empresa",""),
            regional=meta.get("regional",""),uf=meta.get("uf",""),cidade=meta.get("cidade",""),
            stage=stage,close_kind=ck,ev_desc=ev[0] if ev else "",ev_date=ev[1].strftime("%d/%m/%Y") if ev else "",ev_text=ev[2] if ev else "",
            mencionou_orcamento=bool(hit["ORCAMENTO"]),intent_score=iscore,intent_desc=idesc,
            erp_match=has_order,erp_recent=bool(oi.get("data_iso","") and oi.get("data_iso","")>=period_start),
            order_pedido=oi.get("pedido",""),order_status=oi.get("status",""),order_data=oi.get("data",""),
            last=li["last_date"].strftime("%d/%m/%Y"),last_iso=li["last_date"].strftime("%Y-%m-%d"),last_month=li["last_date"].strftime("%Y-%m"),
            client_waiting=bool(client_waiting),seller_replied=bool(seller_replied),contacted_at=(li["last_date"].strftime("%d/%m/%Y") if seller_replied else ""),
            origem="Inbound" if first_dir.get(uuid)=="IN" else "Disparo",
            contact_date=contact_date.strftime("%d/%m/%Y") if contact_date else "",
            contact_date_iso=contact_date.strftime("%Y-%m-%d") if contact_date else "",date_basis=basis))
    return contacts

def build_html(contacts, out_path):
    html=HTML_TEMPLATE.replace("__DATA__", json.dumps(contacts,ensure_ascii=False))
    with open(out_path,"w",encoding="utf-8") as f: f.write(html)


def build_xlsx(contacts, out_path):
    F="Arial"; C=dict(navy="1F3864",erp="0B6B57",done="0E7C66",grey="6B7280")
    thin=Border(*[Side(style="thin",color="D9D9D9")]*4)
    PT={"CONTATAR":"Entrar em contato","ATRASADO":"Atrasado","ORCAMENTO":"Mencionou or\u00e7amento","FECHADO":"Fecharam pedido","PERDIDO":"Perdido","ENTROU":"Entrou em contato","SEM_SINAL":"Sem sinal","STALE":"Pendente (m\u00eas ant.)"}
    COLOR={"CONTATAR":"FBE2CB","ATRASADO":"F8CBCB","ORCAMENTO":"D6E8F5","FECHADO":"C6EFCE","PERDIDO":"E4DCF0","ENTROU":"D6EFE8","SEM_SINAL":"F2F2F2","STALE":"EFEFEF"}
    order={"CONTATAR":0,"ATRASADO":1,"ORCAMENTO":2,"FECHADO":3,"PERDIDO":4,"ENTROU":5,"STALE":6,"SEM_SINAL":7}
    icolor={5:"1E7A4F",4:"C75B12",3:"A8860B"}
    now=datetime.now(); ws_=now.replace(hour=0,minute=0,second=0,microsecond=0)
    week_start=ws_-timedelta(days=ws_.weekday()); month_start=now.replace(day=1,hour=0,minute=0,second=0,microsecond=0)
    def iso(c): return datetime.strptime(c["contact_date_iso"],"%Y-%m-%d") if c["contact_date_iso"] else None
    def fstage(c):
        if c["stage"]=="SEM_SINAL": return "SEM_SINAL"
        if c["stage"]=="CONTATAR":
            d=iso(c)
            if d and d<month_start: return "STALE"
            if d and d<week_start: return "ATRASADO"
            return "CONTATAR"
        return c["stage"]
    wb=Workbook()
    def hrow(ws,r,n,fill):
        for c in range(1,n+1):
            cell=ws.cell(row=r,column=c);cell.font=Font(name=F,bold=True,color="FFFFFF",size=10)
            cell.fill=PatternFill("solid",start_color=fill);cell.alignment=Alignment(vertical="center",wrap_text=True);cell.border=thin
    ws=wb.create_sheet("Funil (conversa)")
    hdr=["Etapa","Inten\u00e7\u00e3o","Contatar at\u00e9","Atendido em","Tem pedido ERP","Cliente","Telefone","UF","Cidade","CNPJ","Empresa","Vendedor","Origem","Or\u00e7amento","\u00daltima Msg","Evid\u00eancia","Contact UUID"]
    ws.append(hdr); hrow(ws,1,len(hdr),C["navy"])
    def sk(c): return (order[fstage(c)],-c["intent_score"],c["contact_date_iso"] or "9999",c["last_iso"])
    for c in sorted(contacts,key=sk):
        es=fstage(c)
        ws.append([PT[es],c["intent_score"] or "",c["contact_date"],c.get("contacted_at",""),
                   ("SIM "+c["order_pedido"]) if c["erp_match"] else "",c["name"],c["urn"],c["uf"],c["cidade"],c["cnpj"],c["empresa"],c["vendedor"],c["origem"],
                   "Sim" if c["mencionou_orcamento"] else "",c["last"],
                   ('%s: "%s"'%(c["ev_desc"],c["ev_text"][:90].strip())) if c["ev_desc"] else "",c["uuid"]])
    for r in ws.iter_rows(min_row=2,max_row=ws.max_row):
        st=[k for k,v in PT.items() if v==r[0].value][0]
        for cell in r: cell.font=Font(name=F,size=9);cell.border=thin
        r[0].fill=PatternFill("solid",start_color=COLOR[st]);r[0].font=Font(name=F,size=9,bold=True)
        iv=r[1].value
        if iv in (3,4,5): r[1].fill=PatternFill("solid",start_color=icolor[iv]);r[1].font=Font(name=F,size=9,bold=True,color="FFFFFF");r[1].alignment=Alignment(horizontal="center")
        if r[4].value: r[4].font=Font(name=F,size=9,bold=True,color="0B6B57")
    for i,w in enumerate([18,9,12,12,16,22,14,6,15,18,22,22,9,9,11,46,38],1): ws.column_dimensions[get_column_letter(i)].width=w
    ws.freeze_panes="A2"; ws.auto_filter.ref="A1:Q%d"%ws.max_row
    we=wb.create_sheet("Confirmado no ERP (CNPJ)")
    h2=["Pedido","Status","Data Pedido","No per\u00edodo","Etapa na conversa","Cliente","CNPJ","Empresa","UF","Vendedor","Telefone","Contact UUID"]
    we.append(h2); hrow(we,1,len(h2),C["erp"])
    erp=[c for c in contacts if c["erp_match"]]
    erp.sort(key=lambda c:(c["order_data"].split("/")[::-1] if c["order_data"] else [""]),reverse=True)
    for c in erp:
        we.append([c["order_pedido"],c["order_status"],c["order_data"],"Sim" if c["erp_recent"] else "",PT.get(c["stage"],c["stage"]),
                   c["name"],c["cnpj"],c["empresa"],c["uf"],c["vendedor"],c["urn"],c["uuid"]])
    for r in we.iter_rows(min_row=2,max_row=we.max_row):
        for cell in r: cell.font=Font(name=F,size=9);cell.border=thin
        r[0].font=Font(name=F,size=9,bold=True,color="0B6B57")
    for i,w in enumerate([20,18,12,10,18,22,18,24,6,22,14,38],1): we.column_dimensions[get_column_letter(i)].width=w
    we.freeze_panes="A2"; we.auto_filter.ref="A1:L%d"%we.max_row
    rs=wb.active;rs.title="Resumo";rs.sheet_view.showGridLines=False
    rs["B2"]="FUNIL COMERCIAL \u2014 AMARA NZERO (Weni x Contatos x Pedidos)"
    rs["B2"].font=Font(name=F,bold=True,size=14,color=C["navy"])
    cnt=collections.Counter(fstage(c) for c in contacts); intent=collections.Counter(c["intent_score"] for c in contacts if c["intent_score"])
    erp_total=sum(1 for c in contacts if c["erp_match"]); erp_rec=sum(1 for c in contacts if c["erp_recent"])
    rows=[("FUNIL PELA CONVERSA",None,C["navy"]),("Entrar em contato",cnt["CONTATAR"]),("Atrasados",cnt["ATRASADO"]),
          ("Mencionou or\u00e7amento",cnt["ORCAMENTO"]),("Fecharam pedido (conversa)",cnt["FECHADO"]),("Perdidos",cnt["PERDIDO"]),
          ("Entrou em contato",cnt["ENTROU"]),("Pendente m\u00eas anterior (oculto)",cnt.get("STALE",0)),("Sem sinal (oculto)",cnt.get("SEM_SINAL",0)),
          ("CONFRONTO COM PEDIDOS (CNPJ)",None,C["erp"]),("Confirmado no ERP (todos)",erp_total),("Confirmado no ERP (no per\u00edodo)",erp_rec),
          ("INTEN\u00c7\u00c3O",None,C["done"]),("Inten\u00e7\u00e3o 5",intent.get(5,0)),("Inten\u00e7\u00e3o 4",intent.get(4,0)),("Inten\u00e7\u00e3o 3",intent.get(3,0))]
    r=5
    for item in rows:
        if len(item)==3:
            rs.cell(row=r,column=2,value=item[0]).font=Font(name=F,bold=True,size=11,color="FFFFFF")
            for c in range(2,5): rs.cell(row=r,column=c).fill=PatternFill("solid",start_color=item[2])
        else:
            rs.cell(row=r,column=2,value=item[0]).font=Font(name=F,size=10)
            vc=rs.cell(row=r,column=4,value=item[1]);vc.font=Font(name=F,bold=True,size=10);vc.alignment=Alignment(horizontal="right")
        r+=1
    rs.column_dimensions["B"].width=42;rs.column_dimensions["C"].width=3;rs.column_dimensions["D"].width=10
    piv=collections.defaultdict(lambda:collections.Counter()); erpby=collections.Counter()
    for c in contacts:
        piv[c["vendedor"]][fstage(c)]+=1
        if c["erp_match"]: erpby[c["vendedor"]]+=1
    wv=wb.create_sheet("Por Vendedor")
    h=["Vendedor","Entrar em contato","Atrasados","Or\u00e7amento","Fecharam (conversa)","Perdidos","Entrou","Confirmado ERP"]
    wv.append(h); hrow(wv,1,len(h),C["done"])
    for vend in sorted(piv,key=lambda v:-(piv[v]["CONTATAR"]+piv[v]["ATRASADO"])):
        p=piv[vend]
        wv.append([vend,p["CONTATAR"],p["ATRASADO"],p["ORCAMENTO"],p["FECHADO"],p["PERDIDO"],p["ENTROU"],erpby.get(vend,0)])
    for rr in wv.iter_rows(min_row=2,max_row=wv.max_row):
        for cell in rr: cell.font=Font(name=F,size=9);cell.border=thin
        rr[0].font=Font(name=F,size=9,bold=True)
        if rr[2].value and rr[2].value>0: rr[2].font=Font(name=F,size=9,bold=True,color="B0322E")
        if rr[7].value and rr[7].value>0: rr[7].font=Font(name=F,size=9,bold=True,color="0B6B57")
    for i,w in enumerate([28,16,11,11,16,10,9,14],1): wv.column_dimensions[get_column_letter(i)].width=w
    wv.freeze_panes="A2"; wv.auto_filter.ref="A1:H%d"%wv.max_row
    wb.save(out_path)

def main():
    ap=argparse.ArgumentParser()
    ap.add_argument("--input",default="./entradas"); ap.add_argument("--output",default="./saidas")
    ap.add_argument("--conversas"); ap.add_argument("--contatos"); ap.add_argument("--pedidos")
    a=ap.parse_args()
    conv=a.conversas or find_file(a.input,"message_export*.xlsx")
    cont=a.contatos or find_file(a.input,"contact_export*.xlsx")
    ped=a.pedidos or find_file(a.input,"*elat*rio*.xlsx","*edido*.xlsx","*rder*.xlsx")
    if not conv: sys.exit("Nao encontrei message_export*.xlsx em %s"%a.input)
    if not cont: sys.exit("Nao encontrei contact_export*.xlsx em %s"%a.input)
    os.makedirs(a.output,exist_ok=True)
    print("[1/4] Pedidos: %s"%(os.path.basename(ped) if ped else "(nenhum - Fecharam pedido ficara vazio)"))
    orders=load_orders(ped)
    print("[2/4] Contatos: %s"%os.path.basename(cont)); ct_map=load_contacts(cont)
    print("[3/4] Conversas: %s"%os.path.basename(conv)); contacts=classify(conv,ct_map,orders)
    print("[4/4] Gerando HTML + XLSX")
    build_html(contacts,os.path.join(a.output,"funil_comercial_amara.html"))
    build_xlsx(contacts,os.path.join(a.output,"funil_comercial_amara.xlsx"))
    cnt=collections.Counter(c["stage"] for c in contacts)
    print("\nOK - %d contatos"%len(contacts))
    erp=sum(1 for c in contacts if c.get("erp_match"))
    print("   Fecharam pedido (conversa): %d"%cnt.get("FECHADO",0))
    print("   Confirmado no ERP (CNPJ):   %d"%erp)
    print("   Entrar em contato:        %d"%cnt.get("CONTATAR",0))
    print("   Entrou em contato:        %d"%cnt.get("ENTROU",0))
    print("   Mencionou orcamento:      %d"%cnt.get("ORCAMENTO",0))
    print("   Perdidos:                 %d"%cnt.get("PERDIDO",0))
    print("\nArquivos em: %s"%os.path.abspath(a.output))
    print("Enviando para o GitHub Pages...")
    upload_to_github(os.path.join(a.output,"funil_comercial_amara.html"),
                     os.environ.get("GITHUB_TOKEN","").strip())

HTML_TEMPLATE = r"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Funil Comercial · Amara NZero · Weni</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link href="https://fonts.googleapis.com/css2?family=Sora:wght@500;600;700&family=Inter:wght@400;500;600&family=IBM+Plex+Mono:wght@500;600&display=swap" rel="stylesheet">
<style>
:root{
  --paper:#EEF1EE;--card:#FFFFFF;--ink:#16243A;--ink-soft:#5A6B82;--line:#DEE3DE;
  --call:#C75B12;--over:#B0322E;--over-bg:#FCF1F0;--orc:#1F6FB2;--orc-bg:#E3EFF8;
  --won:#1E7A4F;--won-bg:#E4F3EA;--erp:#0B6B57;--erp-bg:#DCF1EB;--lost:#5A4A7A;--done:#0E7C66;--done-bg:#E1F2EE;
  --i5:#1E7A4F;--i4:#C75B12;--i3:#A8860B;
}
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:'Inter',sans-serif;background:var(--paper);color:var(--ink);font-size:14px}
header{padding:22px 28px 14px;border-bottom:1px solid var(--line);background:var(--card)}
.eyebrow{font-family:'IBM Plex Mono',monospace;font-size:11px;letter-spacing:.14em;text-transform:uppercase;color:var(--ink-soft)}
h1{font-family:'Sora',sans-serif;font-size:22px;font-weight:700;margin:6px 0 2px}
.sub{color:var(--ink-soft);font-size:13px}
.weeknow{display:inline-block;margin-top:8px;font-family:'IBM Plex Mono',monospace;font-size:11px;background:var(--ink);color:#fff;padding:4px 10px;border-radius:6px}
.totais{display:flex;gap:16px;margin-top:14px;flex-wrap:wrap}
.tot b{font-family:'IBM Plex Mono',monospace;font-size:20px;font-weight:600;display:block}
.tot span{font-size:10.5px;color:var(--ink-soft);text-transform:uppercase;letter-spacing:.05em}
.controls{display:flex;gap:10px;padding:14px 28px;align-items:center;flex-wrap:wrap;background:var(--card);border-bottom:1px solid var(--line)}
.controls input[type=search],.controls select{padding:9px 12px;border:1px solid var(--line);border-radius:8px;font:inherit;background:#fff;color:var(--ink)}
.controls input[type=search]{flex:1;min-width:190px}
.controls input:focus,.controls select:focus{outline:2px solid var(--ink);outline-offset:1px}
.controls label{font-size:11px;color:var(--ink-soft);text-transform:uppercase;letter-spacing:.05em;margin-right:-4px}
.board{display:grid;grid-template-columns:repeat(7,1fr);gap:10px;padding:18px 28px 40px;align-items:start}
@media(max-width:1500px){.board{grid-template-columns:repeat(4,1fr)}}
@media(max-width:1100px){.board{grid-template-columns:repeat(3,1fr)}}
@media(max-width:820px){.board{grid-template-columns:repeat(2,1fr)}}
@media(max-width:600px){.board{grid-template-columns:1fr}.controls,header{padding-left:16px;padding-right:16px}.board{padding:14px 16px 40px}}
.col-head{display:flex;justify-content:space-between;align-items:baseline;padding:9px 4px;position:sticky;top:0;background:var(--paper);z-index:2;border-bottom:2px solid var(--tier)}
.col-head h2{font-family:'Sora',sans-serif;font-size:11px;font-weight:600;text-transform:uppercase;letter-spacing:.02em;color:var(--tier)}
.col-head .count{font-family:'IBM Plex Mono',monospace;font-size:12px;color:var(--ink-soft)}
.col-sep{border-left:2px dashed var(--erp);padding-left:8px}
.card{background:var(--card);border:1px solid var(--line);border-left:3px solid var(--tier);border-radius:10px;padding:11px 13px;margin-top:9px;cursor:pointer;transition:box-shadow .15s}
.card:hover{box-shadow:0 3px 14px rgba(22,36,58,.10)}
.card:focus-visible{outline:2px solid var(--ink);outline-offset:2px}
.cardtop{display:flex;justify-content:space-between;gap:8px;align-items:flex-start}
.nm{font-weight:600;font-size:13px;line-height:1.25}
.tel{font-family:'IBM Plex Mono',monospace;font-size:10.5px;color:var(--ink-soft);margin-top:2px}
.uuid{font-family:'IBM Plex Mono',monospace;font-size:9.5px;color:#9AA6B4;margin-top:2px;word-break:break-all}
.iscore{flex-shrink:0;width:30px;height:30px;border-radius:8px;display:flex;align-items:center;justify-content:center;font-family:'IBM Plex Mono',monospace;font-weight:600;font-size:14px;color:#fff}
.datebox{margin-top:9px;padding:7px 9px;border-radius:7px;background:#F6F8F6;border:1px solid var(--line)}
.datebox .when{font-weight:700;font-size:12.5px;color:var(--tier)}
.datebox .week{font-family:'IBM Plex Mono',monospace;font-size:10.5px;color:var(--ink-soft);margin-top:1px}
.datebox .basis{font-size:10px;color:var(--ink-soft);margin-top:3px;font-style:italic}
.orderbox{margin-top:9px;padding:7px 9px;border-radius:7px;background:var(--erp-bg);border:1px solid #B6DECF}
.orderbox .ped{font-weight:700;font-size:11.5px;color:var(--erp)}
.orderbox .st{font-size:10.5px;color:var(--ink-soft);margin-top:1px}
.badges{display:flex;gap:6px;flex-wrap:wrap;margin-top:8px}
.b{font-size:10.5px;padding:2px 8px;border-radius:999px;background:#EDF0F2;color:var(--ink-soft)}
.b.sell{background:#EAE6F4;color:#5A4A7A}.b.orc{background:var(--orc-bg);color:var(--orc)}.b.inb{background:var(--won-bg);color:var(--won)}
.b.uf{background:#E8EEF4;color:#33506E;font-weight:600}.b.done{background:var(--done-bg);color:var(--done)}.b.erp{background:var(--erp-bg);color:var(--erp);font-weight:600}
.reason{font-size:11.5px;color:var(--ink-soft);margin-top:8px;line-height:1.4}
.evid{display:none;margin-top:10px;padding-top:10px;border-top:1px dashed var(--line);font-size:12.5px;line-height:1.5}
.evid .why{font-weight:600;font-size:10.5px;text-transform:uppercase;letter-spacing:.04em;color:var(--tier);margin-bottom:4px}
.evid q{color:var(--ink-soft);font-style:italic}.evid .extra{margin-top:7px;font-size:11px;color:var(--ink-soft)}
.card.open .evid{display:block}
.more{width:100%;padding:9px;border:1px dashed var(--line);background:transparent;border-radius:10px;cursor:pointer;color:var(--ink-soft);font:inherit;font-size:12px;margin-top:9px}
.empty{color:var(--ink-soft);font-size:12px;padding:12px 4px}
footer{padding:0 28px 34px;color:var(--ink-soft);font-size:11.5px;max-width:1020px;line-height:1.6}
footer b{color:var(--ink)}
.legend{display:flex;gap:14px;flex-wrap:wrap;margin:10px 0 4px}
.legend span{display:flex;align-items:center;gap:5px;font-size:11px}
.dot{width:13px;height:13px;border-radius:4px;display:inline-block}
@media(prefers-reduced-motion:reduce){.card{transition:none}}
</style>
</head>
<body>
<header>
  <div class="eyebrow">Amara NZero · WhatsApp Weni × Contatos × Relatório de Pedidos (CNPJ)</div>
  <h1>Funil Comercial</h1>
  <div class="sub">Funil pela conversa. A última coluna confronta a planilha de pedidos por CNPJ (visão independente).</div>
  <div class="weeknow" id="weeknow"></div>
  <div class="totais" id="totais"></div>
</header>
<div class="controls">
  <input type="search" id="q" placeholder="Buscar nome, telefone, empresa, CNPJ ou cidade…">
  <label>Vendedor</label><select id="f-vend"><option value="">Todos</option></select>
  <label>UF</label><select id="f-uf"><option value="">Todas</option></select>
  <label>Mês</label><select id="f-mes"><option value="">Todos</option></select>
  <label>Origem</label><select id="f-orig"><option value="">Todas</option><option value="Inbound">Inbound</option><option value="Disparo">Disparo</option></select>
</div>
<div class="board" id="board"></div>
<footer>
  <div class="legend">
    <span><i class="dot" style="background:var(--i5)"></i> Intenção 5</span>
    <span><i class="dot" style="background:var(--i4)"></i> Intenção 4</span>
    <span><i class="dot" style="background:var(--i3)"></i> Intenção 3</span>
    <span><i class="dot" style="background:var(--erp)"></i> Tem pedido no ERP (CNPJ)</span>
  </div>
  <strong>Funil pela conversa:</strong> <b>Entrar em contato</b> (quer fechar, aguarda vendedor) · <b>Atrasados</b> (data-alvo vencida no mês vigente) · <b>Mencionou orçamento</b> · <b>Fecharam pedido</b> (a conversa indica fechamento; fluig/repasse não conta) · <b>Perdidos</b> · <b>Entrou em contato</b> (atendente humano já respondeu).
  <br><strong>Bloco separado — <span style="color:var(--erp)">Confirmado no ERP</span>:</strong> confronta o CNPJ com o relatório de pedidos. Lista quem <u>tem pedido real</u> (não cancelado) no relatório, <b>independente do que a conversa disse</b> — pega inclusive quem pediu direto na plataforma sem comentar. Onde houver pedido no ERP, o card mostra a etiqueta <span style="color:var(--erp)">✓ ERP</span> em qualquer coluna.
</footer>
<script>
const DATA=__DATA__;
function startOfWeek(d){const x=new Date(d);const wd=(x.getDay()+6)%7;x.setHours(0,0,0,0);x.setDate(x.getDate()-wd);return x;}
function fmt(d){return String(d.getDate()).padStart(2,'0')+'/'+String(d.getMonth()+1).padStart(2,'0');}
function parseISO(s){const[y,m,dd]=s.split('-').map(Number);return new Date(y,m-1,dd);}
const NOW=new Date();
const WEEK_START=startOfWeek(NOW);const WEEK_END=new Date(WEEK_START);WEEK_END.setDate(WEEK_START.getDate()+6);
const MONTH_START=new Date(NOW.getFullYear(),NOW.getMonth(),1);
document.getElementById('weeknow').textContent='Semana vigente: '+fmt(WEEK_START)+' – '+fmt(WEEK_END)+'  ·  Atrasados a partir de '+fmt(MONTH_START);
function isOverdue(c){if(c.stage!=='CONTATAR'||!c.contact_date_iso)return false;const d=parseISO(c.contact_date_iso);return d<WEEK_START&&d>=MONTH_START;}
function isStale(c){if(c.stage!=='CONTATAR'||!c.contact_date_iso)return false;return parseISO(c.contact_date_iso)<MONTH_START;}
function funnelStage(c){
  if(c.stage==='SEM_SINAL')return 'HIDE';
  if(c.stage==='CONTATAR'){if(isStale(c))return 'HIDE';if(isOverdue(c))return 'ATRASADO';return 'CONTATAR';}
  return c.stage;
}
// colunas do funil (pela conversa) + bloco separado ERP no fim
const STAGES=[
  {id:'CONTATAR',nome:'Entrar em contato',v:'call'},
  {id:'ATRASADO',nome:'Atrasados',v:'over'},
  {id:'ORCAMENTO',nome:'Mencionou orçamento',v:'orc'},
  {id:'FECHADO',nome:'Fecharam pedido',v:'won'},
  {id:'PERDIDO',nome:'Perdidos',v:'lost'},
  {id:'ENTROU',nome:'Entrou em contato',v:'done'},
  {id:'ERP',nome:'✓ Confirmado no ERP',v:'erp',sep:true},
];
const ICOLOR={5:'var(--i5)',4:'var(--i4)',3:'var(--i3)'};
let f={q:'',vend:'',uf:'',mes:'',orig:''};
const shown=Object.fromEntries(STAGES.map(s=>[s.id,(s.id==='ORCAMENTO'||s.id==='ENTROU'||s.id==='ERP')?60:1e9]));
const sv=document.getElementById('f-vend');
[...new Set(DATA.map(c=>c.vendedor))].sort().forEach(v=>{const o=document.createElement('option');o.value=v;o.textContent=v;sv.appendChild(o);});
const suf=document.getElementById('f-uf');
[...new Set(DATA.map(c=>c.uf).filter(Boolean))].sort().forEach(v=>{const o=document.createElement('option');o.value=v;o.textContent=v;suf.appendChild(o);});
const MES_PT={'01':'Jan','02':'Fev','03':'Mar','04':'Abr','05':'Mai','06':'Jun','07':'Jul','08':'Ago','09':'Set','10':'Out','11':'Nov','12':'Dez'};
const sm=document.getElementById('f-mes');
[...new Set(DATA.map(c=>c.last_month))].sort().reverse().forEach(m=>{const o=document.createElement('option');o.value=m;const[y,mm]=m.split('-');o.textContent=MES_PT[mm]+'/'+y;sm.appendChild(o);});
function filtra(){const q=f.q.toLowerCase();
  return DATA.filter(c=>(!f.vend||c.vendedor===f.vend)&&(!f.uf||c.uf===f.uf)&&(!f.mes||c.last_month===f.mes)&&(!f.orig||c.origem===f.orig)&&
    (!q||c.name.toLowerCase().includes(q)||c.urn.includes(q)||(c.empresa||'').toLowerCase().includes(q)||(c.cnpj||'').includes(q)||(c.cidade||'').toLowerCase().includes(q)));}
function esc(s){return String(s).replace(/[&<>"]/g,m=>({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;'}[m]))}
function weekOf(iso){const d=parseISO(iso);const s=startOfWeek(d);const e=new Date(s);e.setDate(s.getDate()+4);return fmt(s)+'–'+fmt(e);}
function card(c,v,isErpCol){
  const d=document.createElement('article');d.className='card';d.tabIndex=0;d.style.setProperty('--tier',`var(--${v})`);
  let dateHtml='';
  const fs=funnelStage(c);
  if(!isErpCol&&(fs==='CONTATAR'||fs==='ATRASADO')&&c.contact_date){
    dateHtml=`<div class="datebox"><div class="when">📞 Contatar até ${esc(c.contact_date)}</div>
      <div class="week">Semana de ${esc(weekOf(c.contact_date_iso))}</div><div class="basis">${esc(c.date_basis)}</div></div>`;
  }
  let orderHtml='';
  if(c.erp_match&&(isErpCol||c.stage==='FECHADO')){
    orderHtml=`<div class="orderbox"><div class="ped">✓ Pedido ${esc(c.order_pedido)}</div><div class="st">${esc(c.order_status)} · ${esc(c.order_data)}</div></div>`;
  }
  let badges=[];
  if(c.uf)badges.push(`<span class="b uf">${esc(c.uf)}</span>`);
  if(c.vendedor)badges.push(`<span class="b sell">${esc(c.vendedor)}</span>`);
  if(c.erp_match&&!isErpCol&&c.stage!=='FECHADO')badges.push(`<span class="b erp">✓ ERP</span>`);
  if(isErpCol)badges.push(`<span class="b">Conversa: ${esc({CONTATAR:'Entrar em contato',ENTROU:'Entrou',ORCAMENTO:'Orçamento',FECHADO:'Fechou',PERDIDO:'Perdido',SEM_SINAL:'Sem sinal'}[c.stage]||c.stage)}</span>`);
  if(c.origem==='Inbound')badges.push(`<span class="b inb">Inbound</span>`);
  if(c.mencionou_orcamento&&c.stage!=='ORCAMENTO'&&!isErpCol)badges.push(`<span class="b orc">Tem orçamento</span>`);
  const iscore=c.intent_score?`<div class="iscore" style="background:${ICOLOR[c.intent_score]}" title="Intenção ${c.intent_score}/5">${c.intent_score}</div>`:'';
  d.innerHTML=`<div class="cardtop"><div><div class="nm">${esc(c.name)}</div>
    ${c.empresa?`<div class="tel">${esc(c.empresa)}</div>`:''}
    <div class="tel">${esc(c.urn)}${c.cidade?' · '+esc(c.cidade):''}${c.uf?'/'+esc(c.uf):''}</div>
    ${c.cnpj?`<div class="tel">${esc(c.cnpj)}</div>`:''}
    <div class="uuid">${esc(c.uuid)}</div></div>${iscore}</div>
    ${dateHtml}${orderHtml}
    <div class="badges">${badges.join('')}</div>
    <div class="reason">últ. msg ${esc(c.last)}${c.client_waiting?' · cliente aguarda':''}</div>
    <div class="evid">${c.ev_desc?`<div class="why">${esc(c.ev_desc)} · ${esc(c.ev_date)}</div><q>“${esc(c.ev_text)}”</q>`:'<q>Sem evidência de fechamento na conversa.</q>'}
      ${c.intent_score?`<div class="extra">Intenção de compra: <b>${c.intent_score}/5</b> — ${esc(c.intent_desc)}</div>`:''}
      <div class="extra">Vendedor: ${esc(c.vendedor)}${c.cnpj?' · CNPJ: '+esc(c.cnpj):''}${c.erp_match?' · Pedido ERP: '+esc(c.order_pedido)+' ('+esc(c.order_status)+')':''}</div></div>`;
  const t=()=>d.classList.toggle('open');d.addEventListener('click',t);
  d.addEventListener('keydown',e=>{if(e.key==='Enter'||e.key===' '){e.preventDefault();t();}});
  return d;
}
function sortFn(s){
  if(s==='ATRASADO')return (a,b)=>(b.intent_score-a.intent_score)||((a.contact_date_iso||'9')>(b.contact_date_iso||'9')?1:-1);
  if(s==='CONTATAR')return (a,b)=>((a.contact_date_iso||'9')>(b.contact_date_iso||'9')?1:-1)||(b.intent_score-a.intent_score);
  if(s==='ERP')return (a,b)=>(b.order_data||'').split('/').reverse().join('').localeCompare((a.order_data||'').split('/').reverse().join(''));
  return (a,b)=>(b.intent_score-a.intent_score)||b.last_iso.localeCompare(a.last_iso);
}
function inColumn(c,s){
  if(s==='ERP') return c.erp_match;          // bloco separado: todo CNPJ com pedido
  return funnelStage(c)===s;                 // funil pela conversa
}
function render(){
  const data=filtra(),board=document.getElementById('board');board.innerHTML='';
  STAGES.forEach(s=>{
    const items=data.filter(c=>inColumn(c,s.id)).sort(sortFn(s.id));
    const col=document.createElement('section');if(s.sep)col.className='col-sep';
    col.innerHTML=`<div class="col-head" style="--tier:var(--${s.v})"><h2>${s.nome}</h2><span class="count">${items.length}</span></div>`;
    if(!items.length)col.insertAdjacentHTML('beforeend','<div class="empty">Nenhum contato.</div>');
    items.slice(0,shown[s.id]).forEach(c=>col.appendChild(card(c,s.v,s.id==='ERP')));
    if(items.length>shown[s.id]){const b=document.createElement('button');b.className='more';
      b.textContent=`Mostrar mais (${items.length-shown[s.id]} restantes)`;b.onclick=()=>{shown[s.id]+=60;render();};col.appendChild(b);}
    board.appendChild(col);
  });
  const t=document.getElementById('totais');
  t.innerHTML=STAGES.map(s=>`<div class="tot"><b style="color:var(--${s.v})">${data.filter(c=>inColumn(c,s.id)).length}</b><span>${s.nome}</span></div>`).join('');
}
['q','f-vend','f-uf','f-mes','f-orig'].forEach(id=>{const el=document.getElementById(id);const key=id==='q'?'q':id.split('-')[1];
  el.addEventListener(id==='q'?'input':'change',e=>{f[key]=e.target.value;render();});});
render();
</script>
</body>
</html>"""



def upload_to_github(html_path, token, repo="Marketing-Amara/Weni", filename="index.html"):
    """Sobe o HTML gerado para o GitHub Pages via API — sem precisar do Git instalado."""
    import base64, urllib.request, urllib.error, json as _json
    if not token:
        print("  GITHUB_TOKEN nao configurado — pulando envio ao GitHub.")
        print("  Configure com:  setx GITHUB_TOKEN \"seu_token\"")
        return
    if not os.path.exists(html_path):
        print("  HTML nao encontrado em %s — pulando." % html_path)
        return
    with open(html_path, "rb") as f:
        content_b64 = base64.b64encode(f.read()).decode("utf-8")
    api_url = "https://api.github.com/repos/%s/contents/%s" % (repo, filename)
    headers = {
        "Authorization": "token %s" % token,
        "Accept": "application/vnd.github+json",
        "Content-Type": "application/json",
        "User-Agent": "funil-weni-script",
    }
    sha = None
    try:
        req = urllib.request.Request(api_url, headers=headers)
        with urllib.request.urlopen(req) as r:
            sha = _json.loads(r.read().decode())["sha"]
    except urllib.error.HTTPError as e:
        if e.code != 404:
            print("  Aviso ao buscar SHA: %d" % e.code)
    body = {"message": "Atualiza funil comercial", "content": content_b64}
    if sha:
        body["sha"] = sha
    try:
        data = _json.dumps(body).encode("utf-8")
        req = urllib.request.Request(api_url, data=data, headers=headers, method="PUT")
        with urllib.request.urlopen(req) as r:
            print("  Publicado: https://marketing-amara.github.io/Weni/")
    except urllib.error.HTTPError as e:
        msg = e.read().decode()
        print("  Erro ao subir para o GitHub (%d): %s" % (e.code, msg[:200]))

if __name__ == "__main__":
    main()
